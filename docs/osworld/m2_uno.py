"""
M2 — GENERAL model authoring via the REAL spreadsheet tool (formulas applied through live LibreOffice/UNO).

The model is given ONLY: the goal, what it observes, and the application's NATIVE operations (set a cell to
a formula/value, add/rename sheets). No library, no evaluator-knowledge, no dtype hacks, no task-specific
hints — those were the test-shaped fixes. The model uses FORMULAS (the spreadsheet's real tool); the live
app computes them (uno_apply.py on the host's LibreOffice). ReAct feedback is GENERAL only (apply error /
file won't load / unchanged). Single tasks may fail honestly; only the BROAD number counts.

HOST-ONLY proxy by default (UNO actually computes, so pandas-vs-gold == compare_table for value tasks).
`--env N` confirms the first N on the real env.evaluate() via M1 reconciliation.
"""
import sys, os, json, glob, re, shutil, subprocess, base64, urllib.request, requests
import pandas as pd

BRAIN = "http://localhost:8080/v1/chat/completions"
EXDIR = "evaluation_examples/examples/libreoffice_calc"
CACHE = "/tmp/m2cache"; os.makedirs(CACHE, exist_ok=True)
WORK = "/tmp/m2_work.xlsx"
OPSF = "/tmp/m2_ops.json"
SYS_PY = "/usr/bin/python3"

PROMPT = (
    "You are operating a LibreOffice Calc spreadsheet to accomplish a task. You act ONLY by issuing "
    "spreadsheet operations; the application computes formulas for you.\n\n"
    "GOAL:\n{instr}\n\n"
    "The open workbook currently contains:\n{struct}\n\n"
    "Operations you can issue (a JSON array, applied in order):\n"
    '  {{"op":"set","sheet":"<name>","cell":"<A1 ref>","formula":"=<Excel-style formula>"}}\n'
    '  {{"op":"set","sheet":"<name>","cell":"<A1 ref>","value":<number or string>}}\n'
    '  {{"op":"add_sheet","name":"<name>","index":<int, 0 = first sheet>}}\n'
    '  {{"op":"rename_sheet","old":"<name>","new":"<name>"}}\n\n'
    "Use formulas for any computation (e.g. =B2-C2, =SUM(F2:H2), cross-sheet =Sheet1!J2). "
    "Respond with ONLY the JSON array."
)


def fetch(url, dest):
    if not os.path.exists(dest):
        urllib.request.urlretrieve(url, dest)
    return dest


def task_io(tf):
    t = json.load(open(tf)); tid = t["id"][:8]
    inurl = inpath = None
    for c in t.get("config", []):
        if c.get("type") == "download":
            f = c["parameters"]["files"][0]; inurl, inpath = f["url"], f["path"]
        elif c.get("type") == "open" and not inpath:
            inpath = c["parameters"]["path"]
    ev = t.get("evaluator", {}); exp = ev.get("expected"); res = ev.get("result")
    if not inurl or not isinstance(exp, dict) or exp.get("type") != "cloud_file":
        return None
    guest_path = (res or {}).get("path") or inpath
    if isinstance(guest_path, (list, tuple)):
        guest_path = guest_path[0]
    if not isinstance(guest_path, str):
        return None
    base = os.path.basename(guest_path)
    try:
        inl = fetch(inurl, "%s/%s_in_%s" % (CACHE, tid, base))
        goldl = fetch(exp["path"], "%s/%s_gold.xlsx" % (CACHE, tid))
    except Exception as e:
        print("   (skip %s: fetch fail %s)" % (tid, e)); return None
    return dict(tid=tid, instr=t["instruction"], inl=inl, goldl=goldl,
                guest_path=guest_path, title=os.path.splitext(base)[0], tf=tf)


def structure(path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(path)
    parts = []
    for ws in wb.worksheets:
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        cols = ", ".join("%s=%r" % (get_column_letter(i + 1), h) for i, h in enumerate(hdr))
        rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
        parts.append("sheet %r: columns(row1)= %s ; data rows 2..%d ; sample: %s"
                     % (ws.title, cols, ws.max_row, rows))
    return "\n".join(parts)


def ask(messages):
    r = requests.post(BRAIN, json={"messages": messages, "temperature": 0, "max_tokens": 2000}, timeout=200)
    txt = r.json()["choices"][0]["message"]["content"]
    m = re.search(r"```(?:json)?\s*(\[.*\])\s*```", txt, re.S) or re.search(r"(\[\s*\{.*\}\s*\])", txt, re.S)
    return (m.group(1) if m else txt).strip(), txt


def apply_ops(ops_text, inl):
    """Returns (ok, why). Parses JSON ops, applies via host UNO, checks file loads + changed."""
    try:
        ops = json.loads(ops_text)
        assert isinstance(ops, list)
    except Exception as e:
        return False, "your output was not a JSON array of operations: %s" % e
    json.dump(ops, open(OPSF, "w"))
    shutil.copy(inl, WORK)
    p = subprocess.run([SYS_PY, "uno_apply.py", WORK, OPSF], capture_output=True, text=True, timeout=120)
    if "APPLIED" not in p.stdout:
        return False, "the spreadsheet engine rejected the operations: %s" % (p.stderr[-400:] or p.stdout[-400:])
    try:
        res = pd.read_excel(WORK, sheet_name=None)
        src = pd.read_excel(inl, sheet_name=None)
        changed = any(sn not in src or not res[sn].equals(src[sn]) for sn in res)
    except Exception as e:
        return False, "result file does not load: %s" % e
    return (changed, "ok" if changed else "the workbook did not change")


def react(t, max_iters=4):
    struct = structure(t["inl"])
    messages = [{"role": "user", "content": PROMPT.format(instr=t["instr"], struct=struct)}]
    for it in range(max_iters):
        ops_text, raw = ask(messages)
        ok, why = apply_ops(ops_text, t["inl"])
        if ok:
            return it
        messages += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": "That did not work: %s\nReturn the corrected JSON array." % why}]
    return max_iters - 1


def predicted_score(goldl):
    try:
        res = pd.read_excel(WORK, sheet_name=None); gold = pd.read_excel(goldl, sheet_name=None)
    except Exception as e:
        return 0.0, "load fail %s" % e
    det, ok = {}, True
    for sn, g in gold.items():
        r = res.get(sn)
        if r is None:
            det[sn] = "missing"; ok = False; continue
        try:
            eq = r.round(4).equals(g.round(4))
        except Exception:
            eq = r.astype(str).reset_index(drop=True).equals(g.astype(str).reset_index(drop=True))
        det[sn] = bool(eq); ok = ok and bool(eq)
    return (1.0 if ok else 0.0), det


# ── broad host-only read (weak proxy — kept only for quick iteration; the real eval is env.evaluate) ──
if __name__ == "__main__":
    N = next((int(a) for a in sys.argv[1:] if a.isdigit()), 12)
    files = sorted(glob.glob(EXDIR + "/*.json"))
    results = []
    print("=== M2 authoring via REAL formula tool (UNO) | host-only over up to %d calc tasks ===" % N, flush=True)
    for tf in files:
        if len(results) >= N:
            break
        t = task_io(tf)
        if not t:
            continue
        try:
            iters = react(t)
            score, det = predicted_score(t["goldl"])
        except Exception as e:
            score, det, iters = 0.0, "EXC %s" % e, -1
        results.append(dict(tid=t["tid"], instr=t["instr"][:60], score=score, iters=iters, detail=det))
        print("  [%s] score=%.0f iters=%s  %s" % (t["tid"], score, iters, t["instr"][:58]), flush=True)
        json.dump(results, open("/tmp/m2_uno_results.json", "w"), indent=1)

    passed = sum(1 for r in results if r["score"] >= 1.0)
    print("\n=== HOST-ONLY (UNO formula tool): %d/%d ===" % (passed, len(results)), flush=True)
    for r in results:
        print("   %s  %.0f  %s" % (r["tid"], r["score"], "" if r["score"] >= 1 else r["detail"]), flush=True)
